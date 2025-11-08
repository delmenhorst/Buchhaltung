from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from pathlib import Path

class ExcelExporter:
    def __init__(self, database):
        self.db = database

    def export_to_excel(self, output_file, filters=None):
        """Export all processed invoices to Excel with income, expenses, and category balances"""
        invoices = self.db.get_all_processed(filters)

        # Sort by date ascending for balance calculations
        invoices_sorted = sorted(invoices, key=lambda x: x['date'] or '')

        # Calculate overall running balance
        balance = 0
        for invoice in invoices_sorted:
            if invoice['invoice_id'] and invoice['invoice_id'].startswith('ERE'):
                balance += invoice['amount'] or 0
            elif invoice['invoice_id'] and invoice['invoice_id'].startswith('ARE'):
                balance -= invoice['amount'] or 0
            invoice['balance'] = balance

        # Sort back to descending (newest first) for export
        invoices_sorted.reverse()

        wb = Workbook()
        ws = wb.active
        ws.title = "EAR-Tabelle"

        # Header row
        headers = ['Datum', 'Kennung', 'Typ', 'Kategorie', 'Beschreibung', 'Einnahme (€)', 'Ausgabe (€)', 'Saldo (€)']
        ws.append(headers)

        # Style header - modern minimalist
        header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        header_font = Font(bold=True, color="212529", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add data
        for invoice in invoices_sorted:
            # Format date as DD.MM.YYYY
            try:
                date_obj = datetime.strptime(invoice['date'], '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d.%m.%Y')
            except:
                formatted_date = invoice['date']

            amount = invoice['amount'] or 0
            invoice_id = invoice['invoice_id'] or ''

            # Determine if income or expense
            is_income = invoice_id.startswith('ERE')
            invoice_type = 'Einnahme' if is_income else 'Ausgabe'

            row = [
                formatted_date,                           # Datum
                invoice_id,                               # Kennung (ERE/ARE)
                invoice_type,                             # Typ
                invoice['category'] or '',                # Kategorie
                invoice['description'] or '',             # Beschreibung
                amount if is_income else 0,               # Einnahme (€)
                amount if not is_income else 0,           # Ausgabe (€)
                invoice['balance']                        # Saldo (running total)
            ]
            ws.append(row)

        # Adjust column widths
        ws.column_dimensions['A'].width = 12   # Datum
        ws.column_dimensions['B'].width = 15   # Kennung
        ws.column_dimensions['C'].width = 12   # Typ
        ws.column_dimensions['D'].width = 15   # Kategorie
        ws.column_dimensions['E'].width = 35   # Beschreibung
        ws.column_dimensions['F'].width = 15   # Einnahme
        ws.column_dimensions['G'].width = 15   # Ausgabe
        ws.column_dimensions['H'].width = 15   # Saldo

        # Format currency columns (F, G, H) - simple format
        for row in range(2, ws.max_row + 1):
            ws[f'F{row}'].number_format = '0.00'
            ws[f'G{row}'].number_format = '0.00'
            ws[f'H{row}'].number_format = '0.00'

        # Add summary row
        ws.append([])
        summary_row = ws.max_row + 1

        total_income = sum(inv['amount'] or 0 for inv in invoices if inv['invoice_id'] and inv['invoice_id'].startswith('ERE'))
        total_expenses = sum(inv['amount'] or 0 for inv in invoices if inv['invoice_id'] and inv['invoice_id'].startswith('ARE'))
        final_balance = total_income - total_expenses

        ws[f'A{summary_row}'] = 'GESAMT'
        ws[f'A{summary_row}'].font = Font(bold=True, size=11)
        ws[f'F{summary_row}'] = total_income
        ws[f'F{summary_row}'].font = Font(bold=True, size=11)
        ws[f'F{summary_row}'].number_format = '0.00'
        ws[f'G{summary_row}'] = total_expenses
        ws[f'G{summary_row}'].font = Font(bold=True, size=11)
        ws[f'G{summary_row}'].number_format = '0.00'
        ws[f'H{summary_row}'] = final_balance
        ws[f'H{summary_row}'].font = Font(bold=True, size=11)
        ws[f'H{summary_row}'].number_format = '0.00'

        # === SHEET 2: Ausgaben nach Kategorien ===
        ws_expenses = wb.create_sheet(title="Ausgaben-Kategorien")

        # Header
        ws_expenses['A1'] = 'Kategorie'
        ws_expenses['B1'] = 'Betrag (€)'
        for cell in ws_expenses[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Calculate expense categories
        expense_categories = {}
        for inv in invoices:
            if inv['invoice_id'] and inv['invoice_id'].startswith('ARE'):
                category = inv.get('category') or 'Ohne Kategorie'
                expense_categories[category] = expense_categories.get(category, 0) + (inv['amount'] or 0)

        # Sort by amount descending
        sorted_expenses = sorted(expense_categories.items(), key=lambda x: x[1], reverse=True)

        # Add data
        row = 2
        for category, amount in sorted_expenses:
            ws_expenses[f'A{row}'] = category
            ws_expenses[f'B{row}'] = amount
            ws_expenses[f'B{row}'].number_format = '0.00'
            row += 1

        # Add total
        row += 1
        ws_expenses[f'A{row}'] = 'GESAMT AUSGABEN'
        ws_expenses[f'A{row}'].font = Font(bold=True, size=11)
        ws_expenses[f'B{row}'] = total_expenses
        ws_expenses[f'B{row}'].font = Font(bold=True, size=11)
        ws_expenses[f'B{row}'].number_format = '0.00'

        # Adjust column widths
        ws_expenses.column_dimensions['A'].width = 25
        ws_expenses.column_dimensions['B'].width = 15

        # === SHEET 3: Einnahmen nach Kategorien ===
        ws_income = wb.create_sheet(title="Einnahmen-Kategorien")

        # Header
        ws_income['A1'] = 'Kategorie'
        ws_income['B1'] = 'Betrag (€)'
        for cell in ws_income[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Calculate income categories
        income_categories = {}
        for inv in invoices:
            if inv['invoice_id'] and inv['invoice_id'].startswith('ERE'):
                category = inv.get('category') or 'Ohne Kategorie'
                income_categories[category] = income_categories.get(category, 0) + (inv['amount'] or 0)

        # Sort by amount descending
        sorted_income = sorted(income_categories.items(), key=lambda x: x[1], reverse=True)

        # Add data
        row = 2
        for category, amount in sorted_income:
            ws_income[f'A{row}'] = category
            ws_income[f'B{row}'] = amount
            ws_income[f'B{row}'].number_format = '0.00'
            row += 1

        # Add total
        row += 1
        ws_income[f'A{row}'] = 'GESAMT EINNAHMEN'
        ws_income[f'A{row}'].font = Font(bold=True, size=11)
        ws_income[f'B{row}'] = total_income
        ws_income[f'B{row}'].font = Font(bold=True, size=11)
        ws_income[f'B{row}'].number_format = '0.00'

        # Adjust column widths
        ws_income.column_dimensions['A'].width = 25
        ws_income.column_dimensions['B'].width = 15

        # Save
        wb.save(output_file)

        return len(invoices)
